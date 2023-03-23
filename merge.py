import pandas
import os
import glob
from pathlib import Path
import openpyxl 
import datetime
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
from change import check_files_for_deploy
from deploy_release import run_process_genfile

class run_process_merge():
    
    def __init__(self, date, storage, container, re_deploy):
        
        self.date = date
        self.re_deploy = re_deploy
        if self.re_deploy == '':
            self.re_deploy = self.date
        
        self.storage = storage
        self.container = container
        self.date_fmt = datetime.datetime.strptime(self.date, '%Y-%m-%d').strftime('%Y%m%d')
        self.str_mvp = ['MVP1','MVP2','MVP3','MVP4','MVP6']
        self.wb = openpyxl.load_workbook("./filename/deployment_checklist_xxxxxxx.xlsx")
        
        # output deployment_checklist
        self.path = os.getcwd() + f'/output/{self.date}'
        os.makedirs(self.path, exist_ok=True)
        self.sheet = self.wb.active
        
        current_path = os.getcwd() + f'/output/{self.date}'
        filename = max(glob.glob(f'{current_path}/template_{self.date_fmt}.xlsx'), key=os.path.getmtime)
        sheet1 = pandas.read_excel(filename, sheet_name='all')
        sheet2 = pandas.read_excel(filename, sheet_name='ddl')
        
        # path
        path_main  = os.getcwd()
        path_table_def = path_main + f'/filename/U02_TABLE_DEFINITION/{self.date}'
        os.makedirs(path_table_def,exist_ok=True)
        path_int_mapp = path_main + f'/filename/U03_INT_MAPPING/{self.date}'
        os.makedirs(path_int_mapp,exist_ok=True)
        path_pl_register = path_main + f'/filename/U99_PL_REGISTER_CONFIG/{self.date}'
        os.makedirs(path_pl_register,exist_ok=True)
        path_pl_ddl = path_main + f'/filename/DDL/{self.date}'
        os.makedirs(path_pl_ddl,exist_ok=True)
        
        df_int_mapp = self.source_int_mapp(sheet1)
        df_sys_name = self.source_pl_config(sheet1)
        df_table_def = self.source_table_def(sheet1)
        df_ddl = self.source_ddl(sheet2)
    
        self.auto_gen_file(df_int_mapp, df_sys_name, df_table_def, df_ddl)
            
    def auto_gen_file(self, df_int_mapp, df_sys_name, df_table_def, df_ddl):
        for mvp in self.str_mvp:
            # config
            sum_df = pandas.concat([df_table_def[mvp], df_int_mapp[mvp], df_sys_name[mvp]], axis=0, ignore_index=True)
            if sum_df.empty is False:
                sum_df['Note UAT Deploy Date'] = self.date
                sum_df['Git_Path'] = sum_df['Storage_Path'].apply(lambda x: "{}/{}/{}/".format(self.date_fmt, x, mvp))
                sum_df["Path"] = sum_df[["Storage_Path", "File_Name"]].apply(lambda x: "/".join(x), axis =1)
                sum_df["Full_Path"] = sum_df[["Git_Path", "File_Name"]].apply(lambda x: "".join(x), axis =1)
                sum_df["Obsolete"] = ""
                sum_df["Checklist"] = sum_df[["Storage", "Container", "Full_Path","Path"]].apply(lambda x: ",".join(x), axis =1)  
                
                df_sum = sum_df.loc[:, ['Storage', 'Container', 'Git_Path', 'Storage_Path','File_Name', 'Obsolete', 'Note UAT Deploy Date', 'Checklist']]
                
                sheet_name = f'Checklist_ADLS_{mvp}'
                self.write_from_source(df_sum, mvp, sheet_name)
                print(f"{mvp} config all row count: {len(df_sum)} rows and write to excel completed.")
            else:
                ("Config folder empty")
                
            # ddl
            if mvp in df_ddl.keys():
                sum_ddl = pandas.DataFrame(df_ddl[mvp])
                sheet_name = f'Checklist_DDL_{mvp}'
                if sum_ddl.empty is False:
                    df_new = sum_ddl.loc[:, ['Storage', 'Container', 'Git_Path', 'Note UAT Deploy Date', 'Checklist']]
                    self.write_from_source(df_new, mvp, sheet_name)
                    print(f"{mvp} ddl row count: {len(df_new)} rows and write to excel completed.")
                else:
                    ("DDL folder empty")
        
    def write_from_source(self, df, mvp, sheet):
        sheet_name = sheet
        self.sheet = self.wb.create_sheet(sheet_name)
        self.sheet.title = sheet_name
        rows = dataframe_to_rows(df, header=True, index=False)
        
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                value = self.sheet.cell(row=r_idx, column=c_idx)
                value.value = val
        
        self.file_name = f'{self.path}/deployment_checklist_{self.date_fmt}.xlsx'
        self.wb.save(self.file_name)
                
    def genarate_datafeame(self, list_df, path):
        dict_df = {}
        for mvp in self.str_mvp:
            df = next((pandas.DataFrame(lst_df[mvp]['File_Name'], columns=['File_Name']).reset_index(drop=True) for lst_df in list_df if mvp in lst_df.keys()), None)
            if df is not None:
                df['Storage'] = self.storage
                df['Container'] = self.container
                df['Storage_Path'] = f'utilities/import/{path}'
                df['MVP'] = mvp
                temp_cols = df.columns.tolist()
                re_cols = temp_cols[1:] + temp_cols[0:1]
                df = df[re_cols]
                
            dict_df.update({mvp: df})
            
        return dict_df
    
    def check_old_deploy(self, sheet):
        current_path = os.getcwd() + r'/filename/OLD_DEPLOY' 
        files_deploy = glob.glob(f'{current_path}/{self.re_deploy}/*')
        if files_deploy != []:
            df_old = pandas.read_excel(''.join(files_deploy), sheet_name=sheet)
        else:
            df_old = pandas.DataFrame()
        
        return df_old
    
    def crate_folder_mvp_and_check_file(self, dataframe, mvp, path):
        
        dict_df = {}
        df_old = self.check_old_deploy(path)
        new_df = dataframe[mvp].apply(lambda x: x.astype(str).str.upper())
        new_df = new_df[~new_df.duplicated(subset=['LIST','MVP'])]
        
        if df_old.empty is False:
            new_df['File_Name'] = pandas.merge(new_df, df_old, how='left', on=['LIST', 'MVP'], indicator=True)
            new_df['File_Name'] = new_df[new_df['_merge'] == 'left_only'].drop('_merge',axis=1)
        
        parent_dir  = f"./filename/{path}/{self.date}"
        destination  = os.path.join(parent_dir, mvp)
        os.makedirs(destination, exist_ok=True)
        
        if path != 'U99_PL_REGISTER_CONFIG':
            new_df['File_Name'] = new_df['LIST'].apply(lambda x: str(x).upper() + '.csv')
        else:
            new_df['File_Name'] = new_df['LIST'].apply(lambda x: f"REGISTER_CONFIG_SYSTEM_{x}.xlsx")
        
        for name in  new_df['File_Name'].values.tolist():
            if os.path.isfile(os.path.join(parent_dir, mvp, name)):
                shutil.copy(os.path.join(parent_dir, name), destination)
            else:
                print(f"Warning!! file: '{name}' does not exist on folder: '{mvp}'")
            
        dict_df.update({mvp: new_df})
        
        return dict_df
        
    def source_int_mapp(self, dataframe):
        
        print("============== int_mapping ==============")
        df = dataframe.loc[dataframe['COL_NM'] == 'INTERFACE_NAME']
        print(f"count row int_mapping: {len(df)} rows")
        
        dict_df = {}
        list_df = []
        if df.empty is False:
            df_mvp = dict(tuple(df.groupby('MVP')))
            
            for mvp in self.str_mvp and df_mvp.keys():
                _df = self.crate_folder_mvp_and_check_file(df_mvp, mvp, path='U03_INT_MAPPING')
                list_df.append(_df)
                
            dict_df = self.genarate_datafeame(list_df, path='U03_INT_MAPPING')
        else:
            [dict_df.update({mvp: pandas.DataFrame()}) for mvp in self.str_mvp]
            
        print("=========================================")
            
        return dict_df
    
    def source_pl_config(self, dataframe):
        
        print("============== pl_config ================")
        df = dataframe.loc[dataframe['COL_NM'] == 'SYSTEM_NAME']
        print(f"count row pl_config: {len(df)} rows")
        
        dict_df = {}
        list_df = []
        if df.empty is False:
            df_mvp = dict(tuple(df.groupby('MVP')))
            
            for mvp in self.str_mvp and df_mvp.keys():
                _df = self.crate_folder_mvp_and_check_file(df_mvp, mvp, path='U99_PL_REGISTER_CONFIG')
                list_df.append(_df)
                
            dict_df = self.genarate_datafeame(list_df, path='U99_PL_REGISTER_CONFIG')
        else:
            [dict_df.update({mvp: pandas.DataFrame()}) for mvp in self.str_mvp]
            
        print("=========================================")
        
        return dict_df
        
    def source_table_def(self, dataframe):
        
        print("============== table_def ================")
        df = dataframe.loc[dataframe['COL_NM'] == 'TABLE']
        new_df = df.loc[~df.duplicated(subset=['LIST','MVP']), :] 
        new_df[['schema', 'table']] = new_df['LIST'].str.split(',', expand=True)
        new_df['LIST'] = new_df['schema'].map(str) + '_' + new_df['table'].map(str)
        print(f"count row table_def: {len(new_df)} rows")
        
        dict_df = {}
        list_df = []
        if new_df.empty is False:
            df_mvp = dict(tuple(new_df.groupby('MVP')))
            
            for mvp in self.str_mvp and df_mvp.keys():
                _df = self.crate_folder_mvp_and_check_file(df_mvp, mvp, path='U02_TABLE_DEFINITION')
                list_df.append(_df)
                
            dict_df = self.genarate_datafeame(list_df, path='U02_TABLE_DEFINITION')
        else:
            [dict_df.update({mvp: pandas.DataFrame()}) for mvp in self.str_mvp]
            
        print("=========================================")
        
        return dict_df
    
    def source_ddl(self, dataframe):
        
        print("================= ddl ===================")
        
        # self.date = self.re_deploy
        ddl_path = os.getcwd() + f'/filename/DDL/{self.date}'
        df = dataframe.loc[~dataframe.duplicated(subset=['VIEW_TABLE','GROUP_JOB_NAME', 'MVP']), :]
        if df.empty is False:
            new_df = df.copy()
            new_df[['Folder','File']] = df['VIEW_TABLE'].map(lambda x: str(x)[2:]).str.split(".", n=1, expand=True)
            print(f"count file ddl: {len(new_df)} files")
            
            df_mvp = dict(tuple(new_df.groupby('MVP')))
            dict_df = {}
            for mvp in self.str_mvp:
                if mvp in df_mvp.keys():
                    destination  = os.path.join(ddl_path, mvp)
                    
                    for fols, files in zip(df_mvp[mvp]["Folder"].values.tolist(), df_mvp[mvp]["File"].values.tolist()):
                        path_in = os.path.join(ddl_path, f'VIEWS/{fols}')
                        path_out =  os.path.join(destination, fols)
                        os.makedirs(path_out, exist_ok=True)
                        full_name = str(files) + '.sql'
                        
                        if os.path.isfile(os.path.join(path_in, full_name)):
                            shutil.copy(os.path.join(path_in, full_name), path_out)
                            check = True
                        else:
                            print(f"Warning!! file: {full_name} does not exist on folder: {fols}, '{mvp}'")
                            check = False
                            
                    # set dataframe
                    sum_df = df_mvp[mvp].reset_index(drop=True)
                    sum_df['File_Name'] = sum_df['File'].apply(lambda x: str(x).upper() + '.sql')
                    sum_df['Note UAT Deploy Date'] = self.date
                    sum_df['Storage'] = self.storage
                    sum_df['Container'] = self.container
                    sum_df['Storage_Path'] = sum_df[["Folder", "File_Name"]].apply(lambda x: "/".join(x), axis=1)
                    sum_df["Git_Path"] =  sum_df['Storage_Path'].apply(lambda x: "VIEWS/{}".format(x)) 
                    sum_df["Checklist"] =  sum_df['Storage_Path'].apply(lambda x: "ddl_script_replace/process_migration/{}".format(x)) 
                    sum_df['MVP'] = mvp
                    dict_df.update({mvp: sum_df})
                    
        if check:
            check_files_for_deploy(self.date, ddl_path=ddl_path)._compare_directories
            
        print("=========================================")
        
        return dict_df
        