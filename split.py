import pandas
import os
import glob
import asyncio
import datetime
import openpyxl 
from openpyxl.utils.dataframe import dataframe_to_rows

class run_process_split():
    
    def __init__(self, date):
        self.date = date
        self.date_fmt = datetime.datetime.strptime(self.date, '%Y-%m-%d').strftime('%Y%m%d')        
        self.wb = openpyxl.Workbook()
        parent_dir  = "./output"
        path = os.path.join(parent_dir, self.date)
        os.makedirs(path, exist_ok=True)
        self.file_name = f'{path}/gen_parameter_{self.date_fmt}.xlsx'
        self.sheet = self.wb.active
        
    def run(self):
        loop = asyncio.get_event_loop()
        loop.run_until_complete(self.find_excel_file())
        
    def join_mvp(self, drop_dup):
        parent_dir  = "./filename"
        ## mvp_file = "SCB and list of GJ and Sys_20230220_update ctl_ID on GJ sheet.xlsx" // old version
        mvp_file = "Status Report to SCB and list of GJ and Sys.xlsx"
        mvp_df = pandas.read_excel(os.path.join(parent_dir, mvp_file), sheet_name='3. Group Job', skiprows=3)
        
        df1 = drop_dup.apply(lambda x: x.astype(str).str.strip().str.upper())
        df2 = mvp_df.apply(lambda x: x.astype(str).str.strip().str.upper())
        m_df = pandas.merge(df1, df2, left_on='GROUP_JOB_NAME', right_on='Group Job', how='left')
        m_df = m_df.drop(m_df.columns[5:], axis=1)
        
        return m_df
    
    def check_period_deploy(self, mvp_drop_dup, period_deploy, sheet):
        
        # old deploy
        date_change = '2023-03-10'
        current_path = os.getcwd() + r'/filename/OLD_DEPLOY' 
        files_deploy = glob.glob(f'{current_path}/{date_change}/*')
        if files_deploy != [] and sheet != 'ddl':
            df_old = pandas.read_excel(''.join(files_deploy), sheet_name=sheet)
        else:
            df_old = pandas.DataFrame()
        
        df1 = mvp_drop_dup.apply(lambda x: x.astype(str).str.upper())
        df2 = period_deploy.apply(lambda x: x.astype(str).str.upper())
        data_all = pandas.merge(df1, df2, left_on='GROUP_JOB_NAME', right_on='GROUP_JOB_NAME', how='left')
        
        if df_old.empty is False:
            data_all = pandas.merge(data_all, df_old, how='left', on=['LIST', 'MVP'], indicator=True)
            data_all = data_all[data_all['_merge'] == 'left_only'].drop('_merge',axis=1)
            
        self.sheet = self.wb.create_sheet(sheet)
        self.sheet.title = sheet
        rows = dataframe_to_rows(data_all, header=True, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                value = self.sheet.cell(row=r_idx, column=c_idx)
                value.value = val
                
        self.wb.save(self.file_name)
        
        return data_all
    
    def write_to_gen_parameter(self, write_list, sheet):
        
        self.sheet = self.wb.create_sheet(sheet)
        self.sheet.title = sheet
        column = ['MVP1', 'MVP2', 'MVP3', 'MVP4', 'MVP6']
        i = 0
        for col in column:
            for dict_data in write_list:
                try:
                    string_val = dict_data[col]
                except:
                    string_val = "[]"
                
                if len(string_val) <= 32767:
                    column = self.sheet.cell(row=i + 1, column=1)
                    column.value = col
                    value = self.sheet.cell(row=i + 2, column=1)
                    value.value = string_val
                    i += 3
                else:
                    txtfile=open(f"./output/{self.date}/{sheet}_{col}.txt", "wt")
                    txtfile.write(string_val)
                    txtfile.close()
            
        self.wb.save(self.file_name)
        
    async def spilt_sheet_ddl(self, dataframe, period_deploy):
        
        print("============== ddl ==============")
        
        # check column duplicate
        mvp_dup = dataframe.loc[dataframe.duplicated(subset=['VIEW_TABLE','GROUP_JOB_NAME','MVP']), :]
        print(f"count duplicated ddl: {len(mvp_dup)}")
        print(f"delete duplicated ddl: {mvp_dup['VIEW_TABLE'].values.tolist()}")
        ## drop duplicate and set to list
        mvp_drop_dup = dataframe[~dataframe.duplicated(subset=['VIEW_TABLE','GROUP_JOB_NAME','MVP'])]
        print(f"count ddl: {len(mvp_drop_dup)}")
        
        # mvp1
        mvp1 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP1']
        mvp1 = mvp1[~mvp1.duplicated(subset=['VIEW_TABLE','MVP'])]
        # mvp2
        mvp2 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP2']
        mvp2 = mvp2[~mvp2.duplicated(subset=['VIEW_TABLE','MVP'])]
        # mvp3
        mvp3 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP3']
        mvp3 = mvp3[~mvp3.duplicated(subset=['VIEW_TABLE','MVP'])]
        # mvp4
        mvp4 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP4']
        mvp4 = mvp4[~mvp4.duplicated(subset=['VIEW_TABLE','MVP'])]
        # mvp6
        mvp6 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP6']
        mvp6 = mvp6[~mvp6.duplicated(subset=['VIEW_TABLE','MVP'])]
        
        # check with deploy from sheet period ./output/template_{date_deploy}.xlsx
        new_df = pandas.concat([mvp1,mvp2,mvp3,mvp4,mvp6], axis=0, ignore_index=True)
        self.check_period_deploy(new_df, period_deploy, sheet='ddl')
        
        ## format output to list
        mvp1_to_list = str(list(mvp1["VIEW_TABLE"])).replace(" ", "\n")
        mvp2_to_list = str(list(mvp2["VIEW_TABLE"])).replace(" ", "\n")
        mvp3_to_list = str(list(mvp3["VIEW_TABLE"])).replace(" ", "\n")
        mvp4_to_list = str(list(mvp4["VIEW_TABLE"])).replace(" ", "\n")
        mvp6_to_list = str(list(mvp6["VIEW_TABLE"])).replace(" ", "\n")
        
        ## format output to list
        write_dict = {}
        write_dict.update({'MVP1': mvp1_to_list,
                        'MVP2': mvp2_to_list,
                        'MVP3': mvp3_to_list,
                        'MVP4': mvp4_to_list,
                        'MVP6': mvp6_to_list,
                        })
        self.write_to_gen_parameter(write_list=[write_dict], sheet="list_ddl")
        
        return 'sheet_ddl completed ..'
    
    async def spilt_table_def(self, dataframe, period_deploy):
        
        print("============== table_def ==============")
        df = dataframe.loc[dataframe['COL_NM'] == 'TABLE']
        
        ## check column duplicate MVP
        mvp_dup = df.loc[df.duplicated(subset=['LIST','GROUP_JOB_NAME','MVP']), :]
        print(f"count duplicated table_definitaion: {len(mvp_dup)}")
        print(f"delete duplicated table_definitaion: {mvp_dup['LIST'].values.tolist()}")
        ## drop duplicate MVP and set to list
        mvp_drop_dup = df[~df.duplicated(subset=['LIST','GROUP_JOB_NAME','MVP'])]
        print(f"count table_definitaion: {len(mvp_drop_dup)}")
        
        # mvp1
        mvp1 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP1']
        mvp1 = mvp1[~mvp1.duplicated(subset=['LIST','MVP'])]
        # mvp2
        mvp2 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP2']
        mvp2 = mvp2[~mvp2.duplicated(subset=['LIST','MVP'])]
        # mvp3
        mvp3 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP3']
        mvp3 = mvp3[~mvp3.duplicated(subset=['LIST','MVP'])]
        # mvp4
        mvp4 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP4']
        mvp4 = mvp4[~mvp4.duplicated(subset=['LIST','MVP'])]
        # mvp6
        mvp6 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP6']
        mvp6 = mvp6[~mvp6.duplicated(subset=['LIST','MVP'])]
            
        # check name duplicate 
        new_df = pandas.concat([mvp1,mvp2,mvp3,mvp4,mvp6], axis=0, ignore_index=True)
        # check with deploy from sheet period ./output/template_{date_deploy}.xlsx
        data_all = self.check_period_deploy(new_df, period_deploy, sheet='U02_TABLE_DEFINITION')
        
        ## format output to list
        write_dict = {}
        for mvp, data in data_all.groupby("MVP"):
            _spilt = data['LIST'].str.split(",", n=1, expand=True)
            schema = str(list(_spilt[0])).replace(" ", "")
            table = str(list(_spilt[1])).replace(" ", "")
            write_dict.update({mvp: f"schema:  {schema}" + ",                 " + f"table:  {table}"})
        
        self.write_to_gen_parameter(write_list=[write_dict], sheet="list_table_definition")
        
        return 'table_def completed ..'
    
    async def spilt_system(self, dataframe, period_deploy):
        
        print("============== system_name ==============")
        df = dataframe.loc[dataframe['COL_NM'] == 'SYSTEM_NAME']
        
        # check column duplicate 
        mvp_dup = df.loc[df.duplicated(subset=['LIST','GROUP_JOB_NAME','MVP']), :]
        print(f"count duplicated system_name: {len(mvp_dup)}")
        print(f"delete duplicated system_name: {mvp_dup['LIST'].values.tolist()}")
        ## drop duplicate and set to list
        mvp_drop_dup = df[~df.duplicated(subset=['LIST','GROUP_JOB_NAME','MVP'])]
        print(f"count system_name: {len(mvp_drop_dup)}")
        
        # mvp1
        mvp1 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP1']
        mvp1 = mvp1[~mvp1.duplicated(subset=['LIST','MVP'])]
        # mvp2
        mvp2 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP2']
        mvp2 = mvp2[~mvp2.duplicated(subset=['LIST','MVP'])]
        # mvp3
        mvp3 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP3']
        mvp3 = mvp3[~mvp3.duplicated(subset=['LIST','MVP'])]
        # mvp4
        mvp4 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP4']
        mvp4 = mvp4[~mvp4.duplicated(subset=['LIST','MVP'])]
        # mvp6
        mvp6 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP6']
        mvp6 = mvp6[~mvp6.duplicated(subset=['LIST','MVP'])]
        
        # check name duplicate 
        new_df = pandas.concat([mvp1,mvp2,mvp3,mvp4,mvp6], axis=0, ignore_index=True)
        # check with deploy from sheet period ./output/template_{date_deploy}.xlsx
        data_all = self.check_period_deploy(new_df, period_deploy, sheet='U99_PL_REGISTER_CONFIG')
        
        ## format output to str
        write_dict = {}
        for mvp, data in data_all.groupby("MVP"):
            write_dict.update({mvp: ','.join(map(str, list(data["LIST"])))})
        self.write_to_gen_parameter(write_list=[write_dict], sheet="list_system_name")
        
        ######## Add REGISTER_CONFIG_SYSTEM_ format ############
        
        new_add_str = pandas.DataFrame(data_all)
        new_add_str['SUFFIX'] = new_add_str['LIST'].apply(lambda x: "{}{}".format('REGISTER_CONFIG_SYSTEM_', x))
        
        new_write_dict = {}
        for mvp, data in new_add_str.groupby("MVP"):
            new_write_dict.update({mvp: str(data['SUFFIX'].values.tolist()).replace(" ", "")})
        self.write_to_gen_parameter(write_list=[new_write_dict], sheet="list_add_suffix_system_name")
        
        return 'system_name completed ..'
    
    async def spilt_int_mapp(self, dataframe, period_deploy):
        
        print("============== int_mapping ==============")
        df = dataframe.loc[dataframe['COL_NM'] == 'INTERFACE_NAME']
        # check column duplicate 
        mvp_dup = df.loc[df.duplicated(subset=['LIST','GROUP_JOB_NAME','MVP']), :]
        print(f"count duplicated int_mapping: {len(mvp_dup)}")
        print(f"delete duplicated int_mapping: {mvp_dup['LIST'].values.tolist()}")
        ## drop duplicate  and set to list
        mvp_drop_dup = df[~df.duplicated(subset=['LIST','GROUP_JOB_NAME','MVP'])]
        print(f"count int_mapping: {len(mvp_drop_dup)}")
        
        # mvp1
        mvp1 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP1']
        mvp1 = mvp1[~mvp1.duplicated(subset=['LIST','MVP'])]
        # mvp2
        mvp2 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP2']
        mvp2 = mvp2[~mvp2.duplicated(subset=['LIST','MVP'])]
        # mvp3
        mvp3 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP3']
        mvp3 = mvp3[~mvp3.duplicated(subset=['LIST','MVP'])]
        # mvp4
        mvp4 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP4']
        mvp4 = mvp4[~mvp4.duplicated(subset=['LIST','MVP'])]
        # mvp6
        mvp6 = mvp_drop_dup[mvp_drop_dup['MVP'] == 'MVP6']
        mvp6 = mvp6[~mvp6.duplicated(subset=['LIST','MVP'])]
        
        # check with deploy from sheet period ./output/template_{date_deploy}.xlsx
        new_df = pandas.concat([mvp1,mvp2,mvp3,mvp4,mvp6], axis=0, ignore_index=True)
        data_all = self.check_period_deploy(new_df, period_deploy, sheet='U03_INT_MAPPING')

        ## format output to list
        write_dict = {}
        for mvp, data in data_all.groupby("MVP"):
            write_dict.update({mvp: str(data['LIST'].values.tolist()).replace(" ", "")})
        self.write_to_gen_parameter(write_list=[write_dict], sheet="list_int_mapping")

        return 'int_mapping completed ..'
        
    async def find_excel_file(self):
        
        current_path = os.getcwd() + '/parameter'
        list_of_files = glob.glob(f'{current_path}/*')
        
        try:
            latest_file = max(list_of_files, key=os.path.getmtime)            
            sheet1 = self.join_mvp(drop_dup=pandas.read_excel(latest_file, sheet_name='all'))
            sheet2 = self.join_mvp(drop_dup=pandas.read_excel(latest_file, sheet_name='ddl'))
            sheet3 = pandas.read_excel(latest_file, sheet_name='period')
            
            with pandas.ExcelWriter(f'./output/{self.date}/template_{self.date_fmt}.xlsx') as writer:
                sheet1.to_excel(writer, sheet_name='all', index=False)
                sheet2.to_excel(writer, sheet_name='ddl', index=False, columns=['VIEW_TABLE', 'GROUP_JOB_NAME', 'MVP'])
                sheet3.to_excel(writer, sheet_name='period', index=False)
            
            coros = [
                asyncio.create_task(self.spilt_int_mapp(sheet1, sheet3)),
                asyncio.create_task(self.spilt_system(sheet1, sheet3)),
                asyncio.create_task(self.spilt_table_def(sheet1, sheet3)),
                asyncio.create_task(self.spilt_sheet_ddl(sheet2, sheet3))
            ]
            results = await asyncio.wait(coros)
            print(f'Completed task: {len(results[0])}')
            [print(f"- {completed_task.result()}") for completed_task in results[0]]    
            print(f'Uncompleted task: {len(results[1])}')
            
        except ValueError as err:
            raise Exception("File not found !!")